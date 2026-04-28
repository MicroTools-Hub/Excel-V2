import base64
import io
import json
import logging
import os
import re
import statistics
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import inspect

from fastapi import FastAPI, HTTPException, Request
from fastapi.concurrency import run_in_threadpool
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from openpyxl import load_workbook
from pydantic import BaseModel, Field

from excel_mcp.calculations import apply_formula as apply_formula_impl
from excel_mcp.chart import create_chart_in_sheet as create_chart_impl
from excel_mcp.data import read_excel_range, write_data
from excel_mcp.formatting import format_range as format_range_impl
from excel_mcp.ollama_client import OllamaClient
from excel_mcp.smart_paste import parse_tabular_text
from excel_mcp.workbook import create_sheet, create_workbook, get_workbook_info
from excel_mcp.pivot import create_pivot_table as create_pivot_table_impl
from excel_mcp.tables import create_excel_table as create_table_impl
from excel_mcp.sheet import (
    copy_range_operation,
    delete_range_operation,
    insert_row,
    insert_cols,
    delete_rows,
    delete_cols,
    delete_sheet,
    merge_range,
    rename_sheet,
    unmerge_range,
)

logger = logging.getLogger("excel-mcp.webapp")

_CELL_REF_RE = re.compile(r"^[A-Za-z]{1,3}[1-9][0-9]{0,6}$")
_RANGE_REF_RE = re.compile(
    r"^[A-Za-z]{1,3}[1-9][0-9]{0,6}(:[A-Za-z]{1,3}[1-9][0-9]{0,6})?$"
)
_SUPPORTED_OPERATION_NAMES = {
    "create_workbook",
    "create_worksheet",
    "rename_worksheet",
    "delete_worksheet",
    "write_data",
    "clear_range",
    "read_data",
    "apply_formula",
    "format_range",
    "create_chart",
    "get_workbook_metadata",
    "create_pivot_table",
    "create_table",
    "insert_rows",
    "insert_columns",
    "delete_sheet_rows",
    "delete_sheet_columns",
    "merge_cells",
    "unmerge_cells",
    "copy_range",
    "delete_range",
}


class ChatMessage(BaseModel):
    role: str
    content: str


class ChatRequest(BaseModel):
    messages: List[ChatMessage]
    filepath: Optional[str] = None
    sheet_name: Optional[str] = None
    start_cell: str = "A1"
    sheet_snapshot: Optional[Dict[str, Any]] = None
    image_context: Optional[Dict[str, Any]] = None
    auto_execute: bool = True


class ToolCallRequest(BaseModel):
    tool_name: str
    args: Dict[str, Any] = Field(default_factory=dict)


class PasteTextRequest(BaseModel):
    filepath: str
    sheet_name: str
    start_cell: str = "A1"
    text: str


class ParseTextRequest(BaseModel):
    text: str


class OcrRequest(BaseModel):
    image_base64: str
    use_ai_layout: bool = False


class OcrPasteRequest(BaseModel):
    filepath: str
    sheet_name: str
    start_cell: str = "A1"
    image_base64: str
    use_ai_layout: bool = False


EXCEL_FILES_BASE = os.path.realpath(os.environ.get("EXCEL_FILES_PATH", "./excel_files"))
os.makedirs(EXCEL_FILES_BASE, exist_ok=True)

APP_ROOT = Path(__file__).resolve().parent
WEB_ROOT = APP_ROOT / "web"
PROJECT_ROOT = APP_ROOT.parent.parent
MANIFEST_TEMPLATE = PROJECT_ROOT / "manifest.xml"

app = FastAPI(
    title="Excel Smart Add-in Backend",
    version="1.0.0",
    description="DeepSeek + Ollama powered backend for Excel add-in workflows.",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

if WEB_ROOT.exists():
    app.mount("/ui", StaticFiles(directory=str(WEB_ROOT), html=True), name="ui")


def _resolved_path_is_within(base: str, candidate: str) -> bool:
    base = os.path.realpath(base)
    candidate = os.path.realpath(candidate)
    if candidate == base:
        return True
    try:
        return os.path.commonpath([base, candidate]) == base
    except ValueError:
        return False


def resolve_excel_path(filename: str) -> str:
    if not filename or "\x00" in filename:
        raise ValueError("Invalid filename")
    if os.path.isabs(filename):
        raise ValueError("filepath must be relative to EXCEL_FILES_PATH")

    candidate = os.path.realpath(os.path.join(EXCEL_FILES_BASE, filename))
    if not _resolved_path_is_within(EXCEL_FILES_BASE, candidate):
        raise ValueError("filepath escapes EXCEL_FILES_PATH")
    return candidate


def _extract_json_payload(text: str) -> Optional[Dict[str, Any]]:
    if not text:
        return None

    tag_match = re.search(r"<excel_plan>(.*?)</excel_plan>", text, flags=re.DOTALL | re.IGNORECASE)
    if tag_match:
        chunk = tag_match.group(1).strip()
        try:
            return json.loads(chunk)
        except json.JSONDecodeError:
            return None

    fenced = re.search(r"```json\s*(.*?)```", text, flags=re.DOTALL | re.IGNORECASE)
    if fenced:
        chunk = fenced.group(1).strip()
        try:
            return json.loads(chunk)
        except json.JSONDecodeError:
            return None

    stripped = text.strip()
    if stripped.startswith("{") and stripped.endswith("}"):
        try:
            return json.loads(stripped)
        except json.JSONDecodeError:
            return None

    return None


def _summarize_sheet(filepath: str, sheet_name: str) -> Dict[str, Any]:
    full_path = resolve_excel_path(filepath)
    workbook_info = get_workbook_info(full_path, include_ranges=False)
    preview = read_excel_range(full_path, sheet_name, start_cell="A1", end_cell="H20")
    return {
        "workbook": workbook_info,
        "preview": preview,
    }


def _compact_sheet_snapshot(snapshot: Dict[str, Any], max_rows: int = 120, max_cols: int = 20) -> Dict[str, Any]:
    def trim_matrix(matrix: Any) -> List[List[Any]]:
        if not isinstance(matrix, list):
            return []
        trimmed: List[List[Any]] = []
        for row in matrix[:max_rows]:
            if isinstance(row, list):
                trimmed.append(row[:max_cols])
            else:
                trimmed.append([row])
        return trimmed

    row_count = int(snapshot.get("row_count") or 0)
    col_count = int(snapshot.get("column_count") or 0)
    truncated = row_count > max_rows or col_count > max_cols

    return {
        "sheet_name": snapshot.get("sheet_name"),
        "sheet_names": snapshot.get("sheet_names", [])[:32] if isinstance(snapshot.get("sheet_names"), list) else [],
        "selection_address": snapshot.get("selection_address"),
        "selection_start_cell": snapshot.get("selection_start_cell"),
        "selection_preview_address": snapshot.get("selection_preview_address"),
        "selection_preview_row_count": snapshot.get("selection_preview_row_count"),
        "selection_preview_column_count": snapshot.get("selection_preview_column_count"),
        "selection_values": trim_matrix(snapshot.get("selection_values", [])),
        "selection_formulas": trim_matrix(snapshot.get("selection_formulas", [])),
        "used_range_address": snapshot.get("used_range_address"),
        "preview_address": snapshot.get("preview_address"),
        "row_count": row_count,
        "column_count": col_count,
        "snapshot_mode": snapshot.get("snapshot_mode") or "compact",
        "truncated": truncated,
        "max_rows": max_rows,
        "max_cols": max_cols,
        "values": trim_matrix(snapshot.get("values", [])),
        "formulas": trim_matrix(snapshot.get("formulas", [])),
    }


def _compact_image_context(image_context: Dict[str, Any], max_rows: int = 80, max_cols: int = 16) -> Dict[str, Any]:
    rows = image_context.get("rows", [])
    normalized_rows = _normalize_table_rows(rows, max_rows=max_rows, max_cols=max_cols)
    text = str(image_context.get("extracted_text") or "").strip()

    return {
        "file_name": image_context.get("file_name"),
        "ocr_engine": image_context.get("ocr_engine"),
        "layout_source": image_context.get("layout_source"),
        "row_count": image_context.get("row_count"),
        "column_count": image_context.get("column_count"),
        "rows": normalized_rows,
        "extracted_text": text[:6000],
    }


def _is_valid_cell_ref(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    token = value.strip()
    if "!" in token:
        token = token.rsplit("!", 1)[-1]
    token = token.replace("$", "").strip()
    return bool(_CELL_REF_RE.fullmatch(token))


def _is_valid_range_ref(value: Any) -> bool:
    if not isinstance(value, str):
        return False
    token = value.strip()
    if "!" in token:
        token = token.rsplit("!", 1)[-1]
    token = token.replace("$", "").replace(" ", "").strip()
    return bool(_RANGE_REF_RE.fullmatch(token))


def _merge_operation_defaults(args: Dict[str, Any], request: ChatRequest) -> Dict[str, Any]:
    def is_missing(value: Any) -> bool:
        if value is None:
            return True
        if isinstance(value, str) and not value.strip():
            return True
        return False

    merged = dict(args)
    if request.filepath and ("filepath" not in merged or is_missing(merged.get("filepath"))):
        merged["filepath"] = request.filepath
    if request.sheet_name and ("sheet_name" not in merged or is_missing(merged.get("sheet_name"))):
        merged["sheet_name"] = request.sheet_name
    if request.start_cell and ("start_cell" not in merged or is_missing(merged.get("start_cell"))):
        merged["start_cell"] = request.start_cell
    return merged


def _validate_operation_pass_one(name: str, args: Dict[str, Any]) -> List[str]:
    errors: List[str] = []

    if not name:
        errors.append("operation name is missing")
        return errors

    if name not in _SUPPORTED_OPERATION_NAMES:
        errors.append(f"unsupported operation '{name}'")
        return errors

    if not isinstance(args, dict):
        errors.append("operation args must be an object")
        return errors

    if name == "write_data" and "data" not in args:
        errors.append("write_data requires data")
    if name == "apply_formula" and "formula" not in args:
        errors.append("apply_formula requires formula")
    if name in {"create_worksheet", "delete_worksheet"} and not args.get("sheet_name"):
        errors.append(f"{name} requires sheet_name")
    if name == "rename_worksheet" and not (args.get("old_name") and args.get("new_name")):
        errors.append("rename_worksheet requires old_name and new_name")

    if name in {"clear_range", "format_range"} and not (
        args.get("range") or args.get("start_cell")
    ):
        errors.append(f"{name} requires range or start_cell")
    if name == "create_chart":
        for required in ("data_range", "chart_type", "target_cell"):
            if not args.get(required):
                errors.append(f"create_chart requires {required}")
    if name == "create_table" and not (args.get("data_range") or args.get("range")):
        errors.append("create_table requires data_range or range")
    pivot_rows = args.get("rows") or args.get("row_fields")
    pivot_values = args.get("values") or args.get("value_fields")
    if name == "create_pivot_table" and not (pivot_rows and pivot_values):
        errors.append("create_pivot_table requires rows and values")

    return errors


def _validate_operation_pass_two(name: str, args: Dict[str, Any], request: ChatRequest) -> List[str]:
    errors: List[str] = []

    if not request.auto_execute and name == "create_workbook":
        errors.append("create_workbook is unavailable in active-workbook mode")

    filepath = args.get("filepath")
    if request.auto_execute and name in {
        "create_workbook",
        "create_worksheet",
        "rename_worksheet",
        "delete_worksheet",
        "write_data",
        "clear_range",
        "read_data",
        "apply_formula",
        "format_range",
        "create_chart",
        "create_pivot_table",
        "create_table",
        "insert_rows",
        "insert_columns",
        "delete_sheet_rows",
        "delete_sheet_columns",
        "merge_cells",
        "unmerge_cells",
        "copy_range",
        "delete_range",
        "get_workbook_metadata",
    }:
        if not isinstance(filepath, str) or not filepath.strip():
            errors.append("server execution requires filepath")
        elif os.path.isabs(filepath):
            errors.append("filepath must be relative, not absolute")

    # In Excel-host mode (auto_execute=False), local execution can default to active sheet.
    if request.auto_execute and name in {
        "create_worksheet",
        "rename_worksheet",
        "delete_worksheet",
        "write_data",
        "clear_range",
        "read_data",
        "apply_formula",
        "format_range",
        "create_chart",
        "create_pivot_table",
        "create_table",
        "insert_rows",
        "insert_columns",
        "delete_sheet_rows",
        "delete_sheet_columns",
        "merge_cells",
        "unmerge_cells",
        "copy_range",
        "delete_range",
    }:
        if not isinstance(args.get("sheet_name"), str) or not str(args.get("sheet_name")).strip():
            errors.append("sheet_name must be a non-empty string")

    if "start_cell" in args and args.get("start_cell") and not _is_valid_cell_ref(args.get("start_cell")):
        errors.append("start_cell is invalid")

    if "end_cell" in args and args.get("end_cell") and not _is_valid_cell_ref(args.get("end_cell")):
        errors.append("end_cell is invalid")

    if "cell" in args and args.get("cell") and not _is_valid_cell_ref(args.get("cell")):
        errors.append("cell is invalid")

    if "range" in args and args.get("range") and not _is_valid_range_ref(args.get("range")):
        errors.append("range is invalid")

    if "data_range" in args and args.get("data_range") and not _is_valid_range_ref(args.get("data_range")):
        errors.append("data_range is invalid")

    if "target_cell" in args and args.get("target_cell") and not _is_valid_cell_ref(args.get("target_cell")):
        errors.append("target_cell is invalid")

    if name == "write_data":
        data = args.get("data")
        if not isinstance(data, list) or not data:
            errors.append("write_data data must be a non-empty list")
        else:
            for row in data:
                if not isinstance(row, list):
                    errors.append("write_data data must be a 2D list")
                    break

    if name == "apply_formula":
        formula = args.get("formula")
        if not isinstance(formula, str) or not formula.strip():
            errors.append("formula must be a non-empty string")

    if name == "clear_range" and not (
        args.get("range") or args.get("start_cell")
    ):
        errors.append("clear_range requires range or start_cell")

    return errors


def _double_check_operations(
    operations: List[Dict[str, Any]],
    request: ChatRequest,
) -> Tuple[List[Dict[str, Any]], List[str], int]:
    accepted: List[Dict[str, Any]] = []
    warnings: List[str] = []
    rejected_count = 0

    for index, op in enumerate(operations[:8], start=1):
        op_name = str(op.get("name", "")).strip().lower()
        op_args = op.get("args") if isinstance(op.get("args"), dict) else {}
        merged_args = _merge_operation_defaults(op_args, request)

        pass_one_errors = _validate_operation_pass_one(op_name, merged_args)
        pass_two_errors = _validate_operation_pass_two(op_name, merged_args, request)

        if pass_one_errors or pass_two_errors:
            rejected_count += 1
            pass_one_text = "; ".join(pass_one_errors) if pass_one_errors else "ok"
            pass_two_text = "; ".join(pass_two_errors) if pass_two_errors else "ok"
            warnings.append(
                f"Operation #{index} '{op_name or 'unknown'}' rejected. "
                f"Check1={pass_one_text}. Check2={pass_two_text}."
            )
            continue

        accepted.append({"name": op_name, "args": merged_args})

    return accepted, warnings, rejected_count


def _execute_operation(name: str, args: Dict[str, Any]) -> Dict[str, Any]:
    operation = name.strip().lower()

    if operation == "create_workbook":
        filepath = resolve_excel_path(args["filepath"])
        create_workbook(filepath)
        return {"ok": True, "message": f"Workbook created: {args['filepath']}"}

    if operation == "create_worksheet":
        filepath = resolve_excel_path(args["filepath"])
        result = create_sheet(filepath, args["sheet_name"])
        return {"ok": True, "message": result.get("message", "Worksheet created")}

    if operation == "rename_worksheet":
        filepath = resolve_excel_path(args["filepath"])
        result = rename_sheet(filepath, args["old_name"], args["new_name"])
        return {"ok": True, "message": result.get("message", "Worksheet renamed")}

    if operation == "delete_worksheet":
        filepath = resolve_excel_path(args["filepath"])
        result = delete_sheet(filepath, args["sheet_name"])
        return {"ok": True, "message": result.get("message", "Worksheet deleted")}

    if operation == "write_data":
        filepath = resolve_excel_path(args["filepath"])
        result = write_data(
            filepath=filepath,
            sheet_name=args.get("sheet_name"),
            data=args["data"],
            start_cell=args.get("start_cell", "A1"),
        )
        return {"ok": True, "message": result.get("message", "Data written")}

    if operation == "read_data":
        filepath = resolve_excel_path(args["filepath"])
        values = read_excel_range(
            filepath=filepath,
            sheet_name=args["sheet_name"],
            start_cell=args.get("start_cell", "A1"),
            end_cell=args.get("end_cell"),
        )
        return {"ok": True, "rows": values}

    if operation == "apply_formula":
        filepath = resolve_excel_path(args["filepath"])
        result = apply_formula_impl(filepath, args["sheet_name"], args["cell"], args["formula"])
        return {"ok": True, "message": result.get("message", "Formula applied")}

    if operation == "format_range":
        filepath = resolve_excel_path(args["filepath"])
        format_range_impl(
            filepath=filepath,
            sheet_name=args["sheet_name"],
            start_cell=args["start_cell"],
            end_cell=args.get("end_cell"),
            bold=args.get("bold", False),
            italic=args.get("italic", False),
            underline=args.get("underline", False),
            font_size=args.get("font_size"),
            font_color=args.get("font_color"),
            bg_color=args.get("bg_color"),
            border_style=args.get("border_style"),
            border_color=args.get("border_color"),
            number_format=args.get("number_format"),
            alignment=args.get("alignment"),
            wrap_text=args.get("wrap_text", False),
            merge_cells=args.get("merge_cells", False),
            protection=args.get("protection"),
            conditional_format=args.get("conditional_format"),
        )
        return {"ok": True, "message": "Range formatted"}

    if operation == "clear_range":
        filepath = resolve_excel_path(args["filepath"])
        if not args.get("sheet_name"):
            raise ValueError("sheet_name is required for clear_range")

        start_cell = str(args.get("start_cell", "A1"))
        end_cell = str(args.get("end_cell", "")).strip()
        range_address = str(args.get("range", "")).strip()
        if not range_address:
            range_address = f"{start_cell}:{end_cell}" if end_cell else start_cell

        wb = load_workbook(filepath)
        if args["sheet_name"] not in wb.sheetnames:
            wb.close()
            raise ValueError(f"Sheet '{args['sheet_name']}' not found")

        ws = wb[args["sheet_name"]]
        cells = ws[range_address]
        if isinstance(cells, tuple):
            for row in cells:
                if isinstance(row, tuple):
                    for cell in row:
                        cell.value = None
                else:
                    row.value = None
        else:
            cells.value = None

        wb.save(filepath)
        wb.close()
        return {"ok": True, "message": f"Cleared range {range_address} in {args['sheet_name']}"}

    if operation == "get_workbook_metadata":
        filepath = resolve_excel_path(args["filepath"])
        info = get_workbook_info(filepath, include_ranges=bool(args.get("include_ranges", False)))
        return {"ok": True, "metadata": info}

    if operation == "create_chart":
        filepath = resolve_excel_path(args["filepath"])
        result = create_chart_impl(
            filepath=filepath,
            sheet_name=args["sheet_name"],
            data_range=args["data_range"],
            chart_type=args.get("chart_type", "bar"),
            target_cell=args.get("target_cell", "H2"),
            title=args.get("title", ""),
            x_axis=args.get("x_axis", ""),
            y_axis=args.get("y_axis", ""),
            style=args.get("style") if isinstance(args.get("style"), dict) else None,
        )
        return {"ok": True, "message": result.get("message", "Chart created")}

    if operation == "create_pivot_table":
        filepath = resolve_excel_path(args["filepath"])
        result = create_pivot_table_impl(
            filepath=filepath,
            sheet_name=args.get("sheet_name"),
            data_range=args.get("data_range"),
            rows=args.get("rows") or args.get("row_fields") or [],
            columns=args.get("columns") or args.get("column_fields") or [],
            values=args.get("values") or args.get("value_fields") or [],
            agg_func=args.get("agg_func", "sum"),
        )
        return {"ok": True, "message": result.get("message", "Pivot table created")}

    if operation == "create_table":
        filepath = resolve_excel_path(args["filepath"])
        data_range = args.get("data_range") or args.get("range")
        if not data_range and args.get("start_cell") and args.get("end_cell"):
            data_range = f"{args['start_cell']}:{args['end_cell']}"
        if not data_range:
            raise ValueError("create_table requires data_range or range")
        result = create_table_impl(
            filepath=filepath,
            sheet_name=args.get("sheet_name"),
            data_range=data_range,
            table_name=args.get("table_name"),
            table_style=args.get("table_style", "TableStyleMedium9"),
        )
        return {"ok": True, "message": result.get("message", "Table created")}

    if operation == "insert_rows":
        filepath = resolve_excel_path(args["filepath"])
        result = insert_row(
            filepath=filepath,
            sheet_name=args.get("sheet_name"),
            start_row=args.get("start_row", 1),
            count=args.get("count", 1),
        )
        return {"ok": True, "message": result.get("message", "Rows inserted")}

    if operation == "insert_columns":
        filepath = resolve_excel_path(args["filepath"])
        result = insert_cols(
            filepath=filepath,
            sheet_name=args.get("sheet_name"),
            start_col=args.get("start_col", 1),
            count=args.get("count", 1),
        )
        return {"ok": True, "message": result.get("message", "Columns inserted")}

    if operation == "delete_sheet_rows":
        filepath = resolve_excel_path(args["filepath"])
        result = delete_rows(
            filepath=filepath,
            sheet_name=args.get("sheet_name"),
            start_row=args.get("start_row", 1),
            count=args.get("count", 1),
        )
        return {"ok": True, "message": result.get("message", "Rows deleted")}

    if operation == "delete_sheet_columns":
        filepath = resolve_excel_path(args["filepath"])
        result = delete_cols(
            filepath=filepath,
            sheet_name=args.get("sheet_name"),
            start_col=args.get("start_col", 1),
            count=args.get("count", 1),
        )
        return {"ok": True, "message": result.get("message", "Columns deleted")}

    if operation == "merge_cells":
        filepath = resolve_excel_path(args["filepath"])
        result = merge_range(
            filepath=filepath,
            sheet_name=args.get("sheet_name"),
            start_cell=args.get("start_cell"),
            end_cell=args.get("end_cell"),
        )
        return {"ok": True, "message": result.get("message", "Cells merged")}

    if operation == "unmerge_cells":
        filepath = resolve_excel_path(args["filepath"])
        result = unmerge_range(
            filepath=filepath,
            sheet_name=args.get("sheet_name"),
            start_cell=args.get("start_cell"),
            end_cell=args.get("end_cell"),
        )
        return {"ok": True, "message": result.get("message", "Cells unmerged")}

    if operation == "copy_range":
        filepath = resolve_excel_path(args["filepath"])
        result = copy_range_operation(
            filepath=filepath,
            sheet_name=args.get("sheet_name"),
            source_start=args.get("source_start"),
            source_end=args.get("source_end"),
            target_start=args.get("target_start"),
            target_sheet=args.get("target_sheet"),
        )
        return {"ok": True, "message": result.get("message", "Range copied")}

    if operation == "delete_range":
        filepath = resolve_excel_path(args["filepath"])
        result = delete_range_operation(
            filepath=filepath,
            sheet_name=args.get("sheet_name"),
            start_cell=args.get("start_cell"),
            end_cell=args.get("end_cell"),
            shift_direction=args.get("shift_direction", "up"),
        )
        return {"ok": True, "message": result.get("message", "Range deleted")}

    # Fallback: try dispatching any existing Excel MCP tool from server.py.
    blocked = {
        "run_sse",
        "run_streamable_http",
        "run_stdio",
        "get_excel_path",
        "_resolved_path_is_within",
    }
    if operation in blocked:
        raise ValueError(f"Unsupported operation: {name}")

    from excel_mcp import server as mcp_server_module

    candidate = getattr(mcp_server_module, operation, None)
    if callable(candidate):
        signature = inspect.signature(candidate)
        accepted_args = {
            key: value
            for key, value in args.items()
            if key in signature.parameters
        }
        previous_base = mcp_server_module.EXCEL_FILES_PATH
        mcp_server_module.EXCEL_FILES_PATH = EXCEL_FILES_BASE
        try:
            output = candidate(**accepted_args)
            return {"ok": True, "message": output}
        finally:
            mcp_server_module.EXCEL_FILES_PATH = previous_base

    raise ValueError(f"Unsupported operation: {name}")


async def _run_llm_chat(request: ChatRequest) -> Dict[str, Any]:
    llm = OllamaClient()
    used_compact_snapshot_fallback = False

    system_content = (
        "You are Excel Copilot running on top of DeepSeek via Ollama. "
        "You are EXTREMELY SMART and understand vague prompts like a human Excel expert. "
        "For every user command, first read the latest active-sheet snapshot that accompanies the request before deciding what to do. "
        "When users ask for calculations (averages, sums, counts, etc.), you MUST: "
        "1. First understand the data structure by examining sheet_snapshot or using read_data if needed "
        "2. Determine appropriate Excel ranges for the calculation "
        "3. Apply correct Excel formulas using apply_formula operation "
        "4. Write results to appropriate cells using write_data if needed "
        ""
        "EXAMPLES OF SMART BEHAVIOR: "
        "- User says 'average of all subjects': Find all subject columns, calculate row-wise averages for each student, write results in new column "
        "- User says 'sum everything': Calculate total sum of all numeric data, place result in logical location "
        "- User says 'format everything properly': Apply consistent formatting (bold headers, borders, number formatting) "
        "- User says 'make it look nice': Auto-fit columns, apply clean formatting, add subtle colors "
        "- User gives short vague prompt: Interpret intent generously and perform complete, polished actions "
        ""
        "When an operation should be executed, include an XML-like plan block exactly in this format: "
        "<excel_plan>{\"assistant_reply\":\"...\",\"operations\":[{\"name\":\"write_data\",\"args\":{...}}]}</excel_plan>. "
        ""
        "Before returning operations, internally verify each operation twice for required args and reference validity. "
        "If auto_execute is false, the user is working in a live Excel workbook, so NEVER use create_workbook. "
        "If the user asks to modify or format data, include at least one actionable operation (for example format_range, write_data, clear_range, or apply_formula). "
        "When live sheet_snapshot context is present, avoid read_data/get_workbook_metadata unless strictly necessary. "
        "If image_context is present, it comes from OCR extracted from a user-provided image. Use it as additional evidence, but treat the live sheet snapshot as the workbook source of truth. "
        ""
        "SUPPORTED OPERATION NAMES: create_workbook, create_worksheet, rename_worksheet, delete_worksheet, write_data, clear_range, read_data, apply_formula, format_range, create_chart, get_workbook_metadata, create_pivot_table, create_table, insert_rows, insert_columns, delete_sheet_rows, delete_sheet_columns, merge_cells, unmerge_cells, copy_range, delete_range. "
        ""
        "CALCULATION PATTERNS: "
        "- For row calculations (averages per student): Use apply_formula with formulas like =AVERAGE(B2:D2) "
        "- For column calculations (totals per subject): Use apply_formula with formulas like =SUM(B2:B10) "
        "- For charts: Use create_chart with {data_range, chart_type, target_cell, title}. "
        "- For native Excel tables: Use create_table with {data_range, table_name, table_style}. "
        "- For complex aggregations: Use create_pivot_table with {data_range, rows, values, columns, agg_func}. "
        "- Always write results to logical locations (next empty column for row calculations, below data for column totals) "
        "- If the sheet already looks like a student marks table, add Total, Average, Percentage, and Grade in adjacent columns for each row and do not create a separate summary block unless explicitly requested. "
        "- For prompts like 'add useful formulas based on the current table', prefer filling formula columns next to the current table rather than creating extra summary sections. "
        ""
        "For write_data, prefer args {start_cell, data, optional sheet_name}; use range only when explicitly requested. "
        "For clearing cells, prefer clear_range with args {range, optional sheet_name}. "
        "If no operation is needed, operations must be an empty list."
        ""
        "BE PROACTIVE: If user asks for something vague like 'make it better' or 'fix this', examine the data and apply appropriate improvements (formatting, calculations, organization)."
    )

    model_messages: List[Dict[str, str]] = [{"role": "system", "content": system_content}]

    if request.filepath and request.sheet_name:
        try:
            context = _summarize_sheet(request.filepath, request.sheet_name)
            model_messages.append(
                {
                    "role": "system",
                    "content": "Workbook context: " + json.dumps(context, default=str),
                }
            )
        except Exception as exc:
            logger.warning("Could not build workbook context: %s", exc)

    if request.sheet_snapshot:
        try:
            snapshot_context = {
                "sheet_name": request.sheet_snapshot.get("sheet_name"),
                "sheet_names": request.sheet_snapshot.get("sheet_names", []),
                "selection_address": request.sheet_snapshot.get("selection_address"),
                "selection_start_cell": request.sheet_snapshot.get("selection_start_cell"),
                "selection_preview_address": request.sheet_snapshot.get("selection_preview_address"),
                "selection_preview_row_count": request.sheet_snapshot.get("selection_preview_row_count"),
                "selection_preview_column_count": request.sheet_snapshot.get("selection_preview_column_count"),
                "selection_values": request.sheet_snapshot.get("selection_values", []),
                "selection_formulas": request.sheet_snapshot.get("selection_formulas", []),
                "used_range_address": request.sheet_snapshot.get("used_range_address"),
                "preview_address": request.sheet_snapshot.get("preview_address"),
                "snapshot_mode": request.sheet_snapshot.get("snapshot_mode"),
                "row_count": request.sheet_snapshot.get("row_count"),
                "column_count": request.sheet_snapshot.get("column_count"),
                "values": request.sheet_snapshot.get("values", []),
                "formulas": request.sheet_snapshot.get("formulas", []),
            }
            model_messages.append(
                {
                    "role": "system",
                    "content": (
                        "Live active-sheet snapshot from Excel host (full used range). "
                        "Treat this as the source of truth for current sheet state: "
                        + json.dumps(snapshot_context, default=str)
                    ),
                }
            )
        except Exception as exc:
            logger.warning("Could not attach sheet snapshot context: %s", exc)

    if request.image_context:
        try:
            compact_image_context = _compact_image_context(request.image_context)
            model_messages.append(
                {
                    "role": "system",
                    "content": (
                        "Latest OCR-derived image context from the user. "
                        "Use this as supporting information for the prompt: "
                        + json.dumps(compact_image_context, default=str)
                    ),
                }
            )
        except Exception as exc:
            logger.warning("Could not attach image context: %s", exc)

    for msg in request.messages[-24:]:
        model_messages.append({"role": msg.role, "content": msg.content})

    try:
        llm_response = await llm.chat(model_messages)
    except Exception as exc:
        message = str(exc)
        timeout_like = "readtimeout" in message.lower() or "timed out" in message.lower()
        if request.sheet_snapshot and timeout_like:
            logger.warning("Full sheet snapshot timed out; retrying with compact snapshot: %s", message)
            compact_snapshot = _compact_sheet_snapshot(request.sheet_snapshot)
            compact_messages = [
                m
                for m in model_messages
                if not (
                    m.get("role") == "system"
                    and isinstance(m.get("content"), str)
                    and m["content"].startswith("Live active-sheet snapshot from Excel host")
                )
            ]
            compact_messages.append(
                {
                    "role": "system",
                    "content": (
                        "Full active-sheet snapshot call timed out upstream. "
                        "Use this compact snapshot to respond quickly and safely: "
                        + json.dumps(compact_snapshot, default=str)
                    ),
                }
            )
            llm_response = await llm.chat(compact_messages)
            used_compact_snapshot_fallback = True
        else:
            raise
    raw_text = llm_response["content"]
    plan = _extract_json_payload(raw_text)

    reply = raw_text.strip()
    operations: List[Dict[str, Any]] = []

    if isinstance(plan, dict):
        if isinstance(plan.get("assistant_reply"), str):
            reply = plan["assistant_reply"]
        ops = plan.get("operations")
        if isinstance(ops, list):
            operations = [op for op in ops if isinstance(op, dict)]

    validated_operations, validation_warnings, rejected_count = _double_check_operations(operations, request)
    operations = validated_operations

    results: List[Dict[str, Any]] = []
    if request.auto_execute and operations:
        for op in operations[:8]:
            op_name = str(op.get("name", "")).strip()
            merged_args = op.get("args") if isinstance(op.get("args"), dict) else {}
            try:
                result = _execute_operation(op_name, merged_args)
                results.append({"name": op_name, "ok": True, "result": result})
            except Exception as exc:
                results.append({"name": op_name, "ok": False, "error": str(exc)})

    return {
        "reply": reply,
        "model": llm_response["model"],
        "operations": operations,
        "operation_results": results,
        "operation_validation": {
            "checked_twice": True,
            "accepted": len(operations),
            "rejected": rejected_count,
        },
        "operation_validation_warnings": validation_warnings,
        "sheet_context_mode": (
            "compact-fallback"
            if used_compact_snapshot_fallback
            else ("full" if bool(request.sheet_snapshot) else "none")
        ),
    }


def _normalize_table_rows(rows: Any, max_rows: int = 600, max_cols: int = 60) -> List[List[Any]]:
    if not isinstance(rows, list):
        return []

    normalized: List[List[Any]] = []
    for row in rows[:max_rows]:
        row_values = row if isinstance(row, list) else [row]
        clean_row: List[Any] = []
        for cell in row_values[:max_cols]:
            if cell is None:
                clean_row.append("")
            elif isinstance(cell, (int, float, bool)):
                clean_row.append(cell)
            else:
                token = str(cell).replace("\r", " ").strip()
                clean_row.append(token[:300])
        if any(str(cell).strip() for cell in clean_row):
            normalized.append(clean_row)

    if not normalized:
        return []

    width = max(len(row) for row in normalized)
    rectangular = [row + [""] * (width - len(row)) for row in normalized]

    last_non_empty_col = -1
    for col in range(width - 1, -1, -1):
        if any(str(row[col]).strip() for row in rectangular):
            last_non_empty_col = col
            break

    if last_non_empty_col < 0:
        return []

    return [row[: last_non_empty_col + 1] for row in rectangular]


def _median(values: List[float], default: float) -> float:
    clean = [float(value) for value in values if isinstance(value, (int, float)) and float(value) > 0]
    return float(statistics.median(clean)) if clean else float(default)


def _decode_base64_image(image_base64: str) -> Any:
    try:
        from PIL import Image, ImageOps
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="OCR dependencies are not installed") from exc

    payload = image_base64.split(",", 1)[1] if "," in image_base64 else image_base64
    try:
        image_bytes = base64.b64decode(payload)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid base64 image payload") from exc

    try:
        image = Image.open(io.BytesIO(image_bytes))
        image = ImageOps.exif_transpose(image)
        image.load()
        return image
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not decode image: {exc}") from exc


def _build_ocr_variants(image: Any) -> List[Tuple[str, Any]]:
    from PIL import Image, ImageEnhance, ImageOps

    base = image.convert("RGB") if image.mode not in {"RGB", "L"} else image.copy()
    lanczos = Image.Resampling.LANCZOS if hasattr(Image, "Resampling") else Image.LANCZOS

    def upscale_if_needed(img: Any) -> Any:
        width, height = img.size
        min_edge = min(width, height)
        if min_edge >= 1200:
            return img
        scale = min(2.5, 1200.0 / max(1, min_edge))
        return img.resize((int(width * scale), int(height * scale)), resample=lanczos)

    upscaled = upscale_if_needed(base)
    gray = ImageOps.grayscale(upscaled)
    contrast = ImageEnhance.Contrast(gray).enhance(2.0)
    binary = contrast.point(lambda p: 255 if p > 165 else 0)

    return [
        ("upscaled", upscaled),
        ("contrast", contrast),
        ("binary", binary),
    ]


def _extract_words_from_ocr_data(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    words: List[Dict[str, Any]] = []
    texts = data.get("text") or []

    for index, raw_text in enumerate(texts):
        token = str(raw_text).strip()
        if not token:
            continue

        def value(name: str, default: Any) -> Any:
            collection = data.get(name) or []
            if index < len(collection):
                return collection[index]
            return default

        try:
            left = int(float(value("left", 0)))
            top = int(float(value("top", 0)))
            width = int(float(value("width", 0)))
            height = int(float(value("height", 0)))
        except Exception:
            continue

        if width <= 0 or height <= 0:
            continue

        conf_raw = value("conf", -1)
        try:
            conf = float(conf_raw)
        except Exception:
            conf = -1.0

        block_num = int(float(value("block_num", 0)))
        par_num = int(float(value("par_num", 0)))
        line_num = int(float(value("line_num", 0)))

        words.append(
            {
                "text": token,
                "left": left,
                "top": top,
                "width": width,
                "height": height,
                "right": left + width,
                "bottom": top + height,
                "center_x": left + (width / 2.0),
                "center_y": top + (height / 2.0),
                "conf": conf,
                "block_num": block_num,
                "par_num": par_num,
                "line_num": line_num,
            }
        )

    return words


def _group_words_into_rows(word_boxes: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not word_boxes:
        return []

    row_tolerance = max(8.0, _median([float(w["height"]) for w in word_boxes], 14.0) * 0.65)
    rows: List[Dict[str, Any]] = []

    for word in sorted(word_boxes, key=lambda w: (float(w["center_y"]), int(w["left"]))):
        best_idx: Optional[int] = None
        best_distance = float("inf")

        for index, row in enumerate(rows):
            distance = abs(float(word["center_y"]) - float(row["center_y"]))
            if distance <= row_tolerance and distance < best_distance:
                best_idx = index
                best_distance = distance

        if best_idx is None:
            rows.append({"center_y": float(word["center_y"]), "words": [word]})
        else:
            row = rows[best_idx]
            row_words = row["words"]
            row_words.append(word)
            row["center_y"] = sum(float(item["center_y"]) for item in row_words) / len(row_words)

    return sorted(rows, key=lambda row: float(row["center_y"]))


def _segment_row_into_cells(row_words: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    if not row_words:
        return []

    words = sorted(row_words, key=lambda w: int(w["left"]))
    widths = [float(w["width"]) for w in words]
    gaps: List[float] = []
    for idx in range(1, len(words)):
        gaps.append(max(0.0, float(words[idx]["left"] - words[idx - 1]["right"])))

    split_threshold = max(
        12.0,
        _median(widths, 18.0) * 0.85,
        _median(gaps, 8.0) * 1.8,
    )

    cells: List[Dict[str, Any]] = []
    active: List[Dict[str, Any]] = [words[0]]
    right_edge = int(words[0]["right"])

    for word in words[1:]:
        gap = int(word["left"]) - right_edge
        if gap > split_threshold:
            text = " ".join(str(item["text"]).strip() for item in active if str(item["text"]).strip()).strip()
            if text:
                left = int(active[0]["left"])
                right = int(max(item["right"] for item in active))
                width = max(1, right - left)
                cells.append(
                    {
                        "text": text,
                        "left": left,
                        "right": right,
                        "width": width,
                        "center_x": left + (width / 2.0),
                    }
                )
            active = [word]
        else:
            active.append(word)
        right_edge = max(right_edge, int(word["right"]))

    text = " ".join(str(item["text"]).strip() for item in active if str(item["text"]).strip()).strip()
    if text:
        left = int(active[0]["left"])
        right = int(max(item["right"] for item in active))
        width = max(1, right - left)
        cells.append(
            {
                "text": text,
                "left": left,
                "right": right,
                "width": width,
                "center_x": left + (width / 2.0),
            }
        )

    return cells


def _table_from_word_boxes(word_boxes: List[Dict[str, Any]], image_width: int) -> List[List[Any]]:
    if not word_boxes:
        return []

    rows = _group_words_into_rows(word_boxes)
    cell_rows = [_segment_row_into_cells(row["words"]) for row in rows]
    all_cells = [cell for row_cells in cell_rows for cell in row_cells if cell.get("text")]

    if not all_cells:
        return []

    median_cell_width = _median([float(cell["width"]) for cell in all_cells], 20.0)
    column_tolerance = max(14.0, median_cell_width * 0.85, float(max(1, image_width)) * 0.035)

    anchors: List[Dict[str, Any]] = []
    for center in sorted(float(cell["center_x"]) for cell in all_cells):
        matched = False
        for anchor in anchors:
            if abs(center - float(anchor["center"])) <= column_tolerance:
                count = int(anchor["count"]) + 1
                anchor["center"] = ((float(anchor["center"]) * int(anchor["count"])) + center) / count
                anchor["count"] = count
                matched = True
                break
        if not matched:
            anchors.append({"center": center, "count": 1})

    merged_anchors: List[Dict[str, Any]] = []
    for anchor in sorted(anchors, key=lambda item: float(item["center"])):
        if not merged_anchors:
            merged_anchors.append(anchor)
            continue
        prev = merged_anchors[-1]
        if abs(float(anchor["center"]) - float(prev["center"])) <= (column_tolerance * 0.5):
            total = int(prev["count"]) + int(anchor["count"])
            prev["center"] = ((float(prev["center"]) * int(prev["count"])) + (float(anchor["center"]) * int(anchor["count"]))) / total
            prev["count"] = total
        else:
            merged_anchors.append(anchor)

    if len(merged_anchors) > 40:
        top = sorted(merged_anchors, key=lambda item: int(item["count"]), reverse=True)[:40]
        merged_anchors = sorted(top, key=lambda item: float(item["center"]))

    if len(merged_anchors) <= 1:
        return []

    table_rows: List[List[Any]] = []
    for row_cells in cell_rows:
        if not row_cells:
            continue

        row_values: List[str] = ["" for _ in merged_anchors]
        for cell in sorted(row_cells, key=lambda item: float(item["center_x"])):
            nearest_idx = min(
                range(len(merged_anchors)),
                key=lambda idx: abs(float(cell["center_x"]) - float(merged_anchors[idx]["center"])),
            )
            text = str(cell.get("text", "")).strip()
            if not text:
                continue
            if row_values[nearest_idx]:
                row_values[nearest_idx] = f"{row_values[nearest_idx]} {text}".strip()
            else:
                row_values[nearest_idx] = text

        if any(value.strip() for value in row_values):
            table_rows.append(row_values)

    return _normalize_table_rows(table_rows)


def _word_boxes_to_text(word_boxes: List[Dict[str, Any]]) -> str:
    lines: List[str] = []
    for row in _group_words_into_rows(word_boxes):
        ordered = sorted(row["words"], key=lambda word: int(word["left"]))
        line = " ".join(str(word.get("text", "")).strip() for word in ordered if str(word.get("text", "")).strip())
        if line:
            lines.append(line)
    return "\n".join(lines).strip()


def _score_table_candidate(rows: List[List[Any]]) -> float:
    if not rows:
        return -1.0
    width = max(len(row) for row in rows)
    non_empty = sum(1 for row in rows for cell in row if str(cell).strip())
    fill_ratio = non_empty / max(1, len(rows) * max(1, width))
    score = (width * 120.0) + float(non_empty) + (250.0 if width > 1 else 0.0) + (fill_ratio * 30.0)
    if len(rows) <= 1:
        score -= 20.0
    return score


def _choose_best_rows(candidates: List[Tuple[str, List[List[Any]]]]) -> Tuple[str, List[List[Any]]]:
    best_source = "none"
    best_rows: List[List[Any]] = []
    best_score = -1.0

    for source, rows in candidates:
        normalized = _normalize_table_rows(rows)
        score = _score_table_candidate(normalized)
        if score > best_score:
            best_score = score
            best_source = source
            best_rows = normalized

    return best_source, best_rows


def _extract_rows_payload(text: str) -> List[List[Any]]:
    payload = _extract_json_payload(text)
    if not isinstance(payload, dict):
        return []
    return _normalize_table_rows(payload.get("rows"))


async def _ai_refine_ocr_rows(
    text: str,
    word_boxes: List[Dict[str, Any]],
    initial_rows: List[List[Any]],
) -> List[List[Any]]:
    lines = [line for line in text.splitlines() if line.strip()]
    if len(lines) < 2:
        return []

    boxes_preview = [
        {
            "text": str(word["text"]),
            "x": int(word["left"]),
            "y": int(word["top"]),
            "w": int(word["width"]),
            "h": int(word["height"]),
            "conf": round(float(word["conf"]), 1),
        }
        for word in word_boxes[:260]
    ]

    system_prompt = (
        "You convert OCR output into clean rectangular Excel tables. "
        "Infer likely columns even when visual grid lines are missing by using x/y box positions and data patterns. "
        "Return strict JSON only, with this exact shape: {\"rows\":[[...],[...]]}. "
        "No markdown, no explanation, no extra keys."
    )
    user_prompt = (
        "Build the best possible table from this OCR output. Preserve row order top-to-bottom. "
        "Use blank cells when a value is missing to keep column alignment.\n\n"
        f"OCR text:\n{text[:12000]}\n\n"
        f"OCR boxes sample (first {len(boxes_preview)} words):\n{json.dumps(boxes_preview, ensure_ascii=True)}\n\n"
        f"Initial heuristic rows:\n{json.dumps(initial_rows[:80], ensure_ascii=True)}"
    )

    try:
        llm = OllamaClient(timeout_seconds=25)
        response = await llm.chat(
            [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            temperature=0.0,
        )
        return _extract_rows_payload(response.get("content", ""))
    except Exception as exc:
        logger.warning("AI OCR refinement failed: %s", exc)
        return []


def _run_tesseract_ocr(image_base64: str) -> Dict[str, Any]:
    try:
        import pytesseract
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="OCR dependencies are not installed") from exc

    image = _decode_base64_image(image_base64)
    variants = _build_ocr_variants(image)
    psm_values = [6, 11]

    best: Optional[Dict[str, Any]] = None
    attempts: List[str] = []

    for variant_name, variant_image in variants:
        for psm in psm_values:
            config = f"--oem 3 --psm {psm} -c preserve_interword_spaces=1"
            try:
                data = pytesseract.image_to_data(variant_image, config=config, output_type=pytesseract.Output.DICT)
            except Exception as exc:
                attempts.append(f"{variant_name}/psm{psm}: {exc}")
                continue

            words = _extract_words_from_ocr_data(data)
            if not words:
                attempts.append(f"{variant_name}/psm{psm}: no words")
                continue

            good_words = sum(1 for word in words if float(word["conf"]) >= 35.0)
            line_keys = {
                (int(word["block_num"]), int(word["par_num"]), int(word["line_num"]))
                for word in words
            }
            score = (good_words * 3.0) + (len(words) * 1.2) + (len(line_keys) * 2.0)

            if best is None or score > float(best["score"]):
                best = {
                    "score": score,
                    "variant": variant_name,
                    "psm": psm,
                    "config": config,
                    "image": variant_image,
                    "word_boxes": words,
                }

    if best is None:
        detail = "OCR processing failed" if not attempts else f"OCR processing failed ({'; '.join(attempts[:4])})"
        raise HTTPException(status_code=400, detail=detail)

    text = ""
    try:
        text = str(pytesseract.image_to_string(best["image"], config=best["config"]))
    except Exception:
        text = ""
    text = text.strip() or _word_boxes_to_text(best["word_boxes"])

    return {
        "text": text,
        "word_boxes": best["word_boxes"],
        "variant": best["variant"],
        "psm": best["psm"],
        "image_width": int(best["image"].size[0]),
        "image_height": int(best["image"].size[1]),
    }


async def _extract_structured_ocr(image_base64: str, use_ai_layout: bool = False) -> Dict[str, Any]:
    ocr = await run_in_threadpool(_run_tesseract_ocr, image_base64)
    text = str(ocr.get("text") or "").strip()
    word_boxes = ocr.get("word_boxes") if isinstance(ocr.get("word_boxes"), list) else []

    text_rows = _normalize_table_rows(parse_tabular_text(text))
    box_rows = _table_from_word_boxes(word_boxes, int(ocr.get("image_width") or 0))
    ai_rows: List[List[Any]] = []

    if use_ai_layout and (len(word_boxes) >= 8 or len(text_rows) >= 3):
        seed_rows = box_rows if box_rows else text_rows
        ai_rows = await _ai_refine_ocr_rows(text, word_boxes, seed_rows)

    source, best_rows = _choose_best_rows(
        [
            ("ai", ai_rows),
            ("geometry", box_rows),
            ("text", text_rows),
        ]
    )

    if not best_rows and text:
        source = "lines"
        best_rows = _normalize_table_rows([[line] for line in text.splitlines() if line.strip()])

    return {
        "text": text,
        "rows": best_rows,
        "row_count": len(best_rows),
        "column_count": max((len(row) for row in best_rows), default=0),
        "line_count": len([line for line in text.splitlines() if line.strip()]),
        "layout_source": source,
        "ocr_engine": "pytesseract",
        "ocr_variant": ocr.get("variant"),
        "ocr_psm": ocr.get("psm"),
        "word_count": len(word_boxes),
    }


@app.get("/health")
def health() -> Dict[str, Any]:
    model = (os.environ.get("OLLAMA_MODEL") or "deepseek-v3.2:cloud").strip()
    base = (os.environ.get("OLLAMA_BASE_URL") or "https://ollama.com").strip()
    return {
        "status": "ok",
        "model": model,
        "ollama_base_url": base,
        "ollama_api_key_configured": bool(os.environ.get("OLLAMA_API_KEY")),
        "excel_files_path": EXCEL_FILES_BASE,
    }


@app.get("/")
def root() -> Any:
    if WEB_ROOT.exists():
        return RedirectResponse(url="/ui/taskpane.html")
    return {"status": "ok", "message": "UI files missing"}


@app.get("/manifest.xml", include_in_schema=False)
def manifest(request: Request) -> Response:
    if not MANIFEST_TEMPLATE.exists():
        raise HTTPException(status_code=404, detail="manifest.xml not found")

    public_base_url = os.environ.get("PUBLIC_BASE_URL")
    if public_base_url:
        base_url = public_base_url.strip().rstrip("/")
    else:
        base_url = str(request.base_url).rstrip("/")

    xml = MANIFEST_TEMPLATE.read_text(encoding="utf-8")
    for placeholder in (
        "https://excel-v2.onrender.com",
        "https://smart-excel-copilot.onrender.com",
        "https://your-render-service.onrender.com",
        "https://localhost:3000",
    ):
        xml = xml.replace(placeholder, base_url)
    return Response(content=xml, media_type="application/xml")


@app.get("/ui/taskpane.html", include_in_schema=False)
def taskpane() -> Any:
    path = WEB_ROOT / "taskpane.html"
    if not path.exists():
        raise HTTPException(status_code=404, detail="taskpane.html not found")
    return FileResponse(path)


@app.post("/api/chat")
async def chat(request: ChatRequest) -> Dict[str, Any]:
    if not request.messages:
        raise HTTPException(status_code=400, detail="messages cannot be empty")
    try:
        return await _run_llm_chat(request)
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("Chat request failed: %s", exc)
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@app.post("/api/tool-call")
def tool_call(request: ToolCallRequest) -> Dict[str, Any]:
    try:
        result = _execute_operation(request.tool_name, request.args)
        return {"ok": True, "result": result}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/parse-text")
def parse_text_endpoint(request: ParseTextRequest) -> Dict[str, Any]:
    rows = parse_tabular_text(request.text)
    return {
        "rows": rows,
        "row_count": len(rows),
        "column_count": max((len(r) for r in rows), default=0),
    }


@app.post("/api/paste-text")
def paste_text(request: PasteTextRequest) -> Dict[str, Any]:
    rows = parse_tabular_text(request.text)
    if not rows:
        raise HTTPException(status_code=400, detail="No tabular content detected")

    full_path = resolve_excel_path(request.filepath)
    result = write_data(
        filepath=full_path,
        sheet_name=request.sheet_name,
        data=rows,
        start_cell=request.start_cell,
    )
    return {
        "ok": True,
        "message": result.get("message", "Data written"),
        "rows_written": len(rows),
        "columns_written": max((len(r) for r in rows), default=0),
    }


@app.post("/api/ocr-text")
async def ocr_text(request: OcrRequest) -> Dict[str, Any]:
    return await _extract_structured_ocr(
        image_base64=request.image_base64,
        use_ai_layout=bool(request.use_ai_layout),
    )


@app.post("/api/paste-ocr")
async def paste_ocr(request: OcrPasteRequest) -> Dict[str, Any]:
    ocr_payload = await _extract_structured_ocr(
        image_base64=request.image_base64,
        use_ai_layout=bool(request.use_ai_layout),
    )
    text = str(ocr_payload.get("text") or "")
    rows = ocr_payload.get("rows") if isinstance(ocr_payload.get("rows"), list) else []
    if not rows:
        raise HTTPException(status_code=400, detail="OCR succeeded but no tabular data was detected")

    full_path = resolve_excel_path(request.filepath)
    result = write_data(
        filepath=full_path,
        sheet_name=request.sheet_name,
        data=rows,
        start_cell=request.start_cell,
    )
    return {
        "ok": True,
        "message": result.get("message", "Data written"),
        "ocr_text": text,
        "layout_source": ocr_payload.get("layout_source"),
        "ocr_variant": ocr_payload.get("ocr_variant"),
        "ocr_psm": ocr_payload.get("ocr_psm"),
        "rows_written": len(rows),
        "columns_written": max((len(r) for r in rows), default=0),
    }


def run_web_app() -> None:
    import uvicorn

    host = os.environ.get("WEBAPP_HOST", "0.0.0.0")
    port = int(os.environ.get("PORT", "10000"))
    uvicorn.run("excel_mcp.webapp:app", host=host, port=port)


if __name__ == "__main__":
    run_web_app()
